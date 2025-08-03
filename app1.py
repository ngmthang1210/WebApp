from flask import Flask, request, send_file, render_template_string
from openpyxl import load_workbook
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
import os

app = Flask(__name__)

#UPLOAD_FOLDER = 'uploads'
RESULT_FOLDER = 'ToPrint'
#os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(RESULT_FOLDER, exist_ok=True)

HTML_FORM = """
<!DOCTYPE html>
<title>Excel In Tool</title>
<h2>Upload file Excel 01 để chuyển sang 01P (in)</h2>
<form method=post enctype=multipart/form-data>
  <input type=file name=file accept=".xlsx">
  <input type=submit value=Chuyển>
</form>
{% if filename %}
  <p>Đã xử lý xong: <a href="{{ url_for('download_file', filename=filename) }}">Tải file kết quả</a></p>
{% endif %}
"""
@app.route('/', methods=['GET', 'POST'])
def upload_file():
    filename = None
    if request.method == 'POST':
        f = request.files['file']
        if not f or not f.filename.lower().endswith('.xlsx'):
            return "Chỉ nhận file .xlsx!", 400
        in_path = os.path.join(UPLOAD_FOLDER, f.filename)
        out_filename = f.filename[:-5] + 'P.xlsx'
        out_path = os.path.join(RESULT_FOLDER, out_filename)
        f.save(in_path)
        process_excel(in_path, out_path)
        filename = out_filename
    return render_template_string(HTML_FORM, filename=filename)
@app.route('/download/<filename>')
def download_file(filename):
    out_path = os.path.join(RESULT_FOLDER, filename)
    abs_path = os.path.abspath(out_path)  # Convert to absolute path
    if not os.path.exists(abs_path):
        return f"File not found! File xử lý chưa được tạo ra.<br>Đường dẫn tìm: {abs_path}", 404
    return send_file(abs_path, as_attachment=True)

def process_excel(input_file, output_file):
    wb = load_workbook(input_file)
    ws = wb.active
  # Lấy serial trước khi xóa cột!
    serial_val = ws['I2'].value #if 'I2' in ws and ws['I2'].value else ''
    ws.insert_rows(1, amount=1)
    ws['A1'].value = "KV.TB.CB.01"
    ws['B1'].value = "Kho NL"
     # Xóa các cột
    ws.delete_cols(5)
    ws.delete_cols(4)
    ws.delete_cols(8)
    ws.delete_cols(7)
    ws.delete_cols(6)
    ws.delete_cols(5)
    # Gán lại serial vào C1, rồi merge cell
    ws['C1'].value = 'Serial: ' + str(serial_val)
    ws.merge_cells('C1:D1')

    ws.page_margins = PageMargins(left=0.3, right=0.3, top=0.4, bottom=0.4)
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    for idx, col in enumerate(ws.columns, 1):
        max_length = 0
        col_letter = get_column_letter(idx)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 15)

    wb.save(output_file)
    print(f"Đã lưu file: {output_file}")

if __name__ == '__main__':
    app.run(debug=True)
