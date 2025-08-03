from openpyxl import load_workbook
from openpyxl.worksheet.page import PageMargins
from openpyxl import load_workbook
from openpyxl.worksheet.page import PageMargins

# 1. Mở file gốc
wb = load_workbook("fr1006_1606cb01.xlsx")
ws = wb.active
ws.insert_rows(1, amount=1)  # Chèn 2 dòng trống trên cùng
# 3. Merge ô tiêu đề, ví dụ A1 tới D1       ws.merge_cells('A1:D1')
ws['A1'].value = "KV.TB.CB.01"
ws['B2'].value = "Kho NL"
#ws['A1'].alignment = ws['A1'].alignment.copy(horizontal="center")
# 4. Xóa cột (ví dụ xóa cột thứ 5 và 2) - phải xóa từ cột có số lớn xuống nhỏ
ws.delete_cols(5)  # Xóa cột E
ws.delete_cols(4)  # Xóa cột B (sau khi đã xóa cột E, thứ tự đã thay đổi)
ws['C1'].value = 'Serial: ' +  ws['G3'].value
ws.delete_cols(8)  # Xóa cột E
ws.delete_cols(7)
ws.delete_cols(6)  # Xóa cột E
ws.delete_cols(5)
ws.merge_cells('C1:D1')  
# 5. Căn lề, A4 dọc, fit to width
ws.page_margins = PageMargins(left=0.3, right=0.3, top=0.4, bottom=0.4)
ws.page_setup.orientation = "portrait"
ws.page_setup.paperSize = ws.PAPERSIZE_A4
ws.page_setup.fitToPage = True
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
# 6. Điều chỉnh chiều rộng từng cột hợp lý (max 15 ký tự)
from openpyxl.utils import get_column_letter
for idx, col in enumerate(ws.columns, 1):
    max_length = 0
    col_letter = get_column_letter(idx)
    for cell in col:
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))
    ws.column_dimensions[col_letter].width = min(max_length + 2, 15)

# 7. Lưu lại file mới
wb.save("fr1006_1606cb01P.xlsx")
